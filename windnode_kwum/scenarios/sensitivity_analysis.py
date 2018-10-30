import itertools
import numpy as np
from windnode_kwum.scenarios import reference_scenario_curtailment
import pandas as pd
import matplotlib.pyplot as plt
import os

# import the tool that modifies the excel sheet. If a problem is shown here, you have to go to:
# PyCharm -> File -> Default Settings -> Project interpreter python 3.6 -> + (plus symbol) -> openpyxl-> install package
from openpyxl import load_workbook

# import csv to store the data results. If a problem is shown here, you have to go to:
# PyCharm -> File -> Default Settings -> Project interpreter python 3.6 -> + (plus symbol) -> cvs -> install package
import csv

# import seaborn to plot the heatmap. If a problem is shown here, you have to go to:
# PyCharm -> File -> Default Settings -> Project interpreter python 3.6 -> + (plus symbol) -> seaborn -> install package
import seaborn as sns; sns.set()

# import SALib to do sensitivity analysis with the sobol method. If a problem is shown here, you have to go to:
# PyCharm -> File -> Default Settings -> Project interpreter python 3.6 -> + (plus symbol) -> SALib -> install package
from SALib.analyze import sobol


##### HOW THIS PROGRAM WORKS? #####

# This sensitivity analysis works changing two different parameters already existing in the excel file reference_scenario_curtailment.xls
# This program do the sensitivity analysis as follows:

# 1) SAMPLING GENERATION:
    #  1.1) The program will open the excel file reference_scenario_curtailment.xlsx and will change the two desired parameters
    #  1.2) The changes of the excel file will be saved
    #  1.3) The reference_scenario_curtailmente.py will be executed n-run times to get the results of the optimization of the energy system
    #  1.4) The results of the optimization and the combination of the parameters will be saved on a csv file called sensitivity_results.csv
    #     located at /.../WindNODE_KWUM/windnode_kwum/scenarios/results
    #  1.5) A heat map will be plot using the data stored on the file sensitivy_results.csv

# 2) SENSITIVITY ANALYSIS USING THE SOBOL METHDOLOGY WITH SALIB
    # ABOUT SALIB: https://salib.readthedocs.io/en/latest/basics.html
    # ABOUT SOBOL SENSITIVITY ANALYSIS: https://salib.readthedocs.io/en/latest/api.html#sobol-sensitivity-analysis
        # 2.1) The resutls stored on the file sensitivy_results.csv  will be used as the outputs required by the SOBOL-Sensitivity
        #      analysis
        # 2.2) The sobol sensitivity analysis will be shown at the end of the run



##### DEFINITION OF PARAMETERS AND VALUES FOR SAMPLING #####

# function to create combinations of parameter values while preserving parameter names
def product_dict(**kwargs):
    keys = kwargs.keys()
    values = kwargs.values()
    for instance in itertools.product(*values):
        yield dict(zip(keys, instance))

# Sensitive analysis works changing two different parameters already existing in the excel file reference_scenario_curtailment.xls
# Define the variables of the parameters to be varied.
# x_param -> x axis for plotting purposes, y_param -> y axis for plotting purposes,
# Format of the name: 'tab.row.column'
x_param_to_be_varied = "transformers.P2H_sch.capacity"
y_param_to_be_varied = "storages.storage_th_sch.nominal capacity"
# From the flow results of the energy optimization (bus_results), get the flow value needed for the sensitivity analysis.
# The format is:   "from_bus_name.to_component_name"
value_to_extract = "bus_flex.P2H_pr"

# Format for the range and step [min, max, step]
params_to_be_varied = {y_param_to_be_varied: [1, 120, 40], x_param_to_be_varied:[1, 12, 4]}

# create ranges
param_val_ranges = {}
for key, val in params_to_be_varied.items():
    param_val_ranges[key] = list(np.arange(val[0], val[1]+val[2], val[2]))

# create combinations using those ranges
param_val_combinations = list(product_dict(**param_val_ranges))

# Find the current file path for sensitivity_analysis.py
dirpath = os.getcwd()

# Create a csv file containing the data of the sensitivity analysis. This file is called sensitivity_results.csv
# and is stored  at: /.../WindNODE_KWUM/windnode_kwum/scenarios/results
# csvData = [['Variable 1', 'Variable 2', "result"]]
csvData = [[y_param_to_be_varied, x_param_to_be_varied, "result"]]

# Start the Combination of parameters_to_be_varied
for run_no, comb in enumerate(param_val_combinations):
    print('Run no ', str(run_no))
    print('=============')
    parameter_set = []
    data_from_run = []
    for key, val in comb.items():
        print('Parameter', key, '=', val)

        parameter_set.append({'key' : key, 'value': str(val)})

        data_from_run.append(str(val))


    # scenario_file = os.path.join('reference_scenario_curtailment.xlsx')
    dest = dirpath+'/data/reference_scenario_curtailment.xlsx'


##### OPEN, MODIFY AND SAVE EXCEL FILE #####

    # Open the reference_scenario_curtailment.xlsx for reading
    wb = load_workbook(filename=dest)

    for param in parameter_set:

        file_sections = param.get('key').split('.')

        # Open the sheet specified on the parameter_to_be_varied -> (tab.row.column)
        ws = wb[file_sections[0]] #wb.get_sheet_by_name(file_sections[0])

        # Edit the cell specified on the parameter_to_be_varied -> (tab.row.column)
        row_to_edit = ''
        column_to_edit = ''

        for cell in ws.get_cell_collection():
            if cell.value == file_sections[1]:
                row_to_edit = cell.row

            if cell.value == file_sections[2]:
                column_to_edit = cell.col_idx

        ws.cell(row=row_to_edit, column=column_to_edit).value = param.get('value')

    # Save excel file with the changes made
    wb.save(dest)


##### EXECUTE reference_scenario_curtailment AND EXTRACT THE VALUE NEEDED from the energy optimization results:bus_results #####

    # Run the reference_scenario_curtailment.py to execute the  optimization of the energy system
    SA_variables = {
        "is_active": True,
        "results": data_from_run,
        "value_to_extract": value_to_extract
    }

    reference_scenario_curtailment.executeMain(SA_variables=SA_variables)

    csvData.append(data_from_run)

    print('======== Run END ========')


##### SAVE THE RESULTS OF THE SENSITIVITY ANALYSIS #####

# Write sensitivity analysis results in a CSV file called sensitivity_results.csv located in: WindNODE_KWUM->scenarios->results
with open(dirpath+'/results/sensitivity_results.csv', 'w') as csvFile:
    writer = csv.writer(csvFile)
    writer.writerows(csvData)

csvFile.close()


##### PLOT RESULTS OF THE SENSITIVITY ANALYSIS IN A HEATMAP FORMAT (SEABORN) #####


#Plot heatmap of with the results of each run of the energy system optimization
sensitivity_heatmap_file = pd.read_csv(dirpath+"/results/sensitivity_results.csv")
sensitivity_heatmap = sensitivity_heatmap_file.pivot(y_param_to_be_varied, x_param_to_be_varied, "result")
# The palette color of heatmap can be changed with the following options:
# cmap="YlGnBu", cmap="Blues", cmap="BuPu", cmap="Greens"
ax = sns.heatmap(sensitivity_heatmap, cmap="Greens", annot=True, fmt=".1f")
ax.invert_yaxis()
plt.show()