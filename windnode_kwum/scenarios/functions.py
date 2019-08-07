# define and setup logger
import oemof
import numpy as np
import glob
import os
import win32com.client as win32
import shutil
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import glob
import os
import win32com.client as win32
import shutil
from openpyxl import load_workbook
from win32com.client import DispatchEx
from windnode_kwum.models.basic_model import create_model, simulate

# define and setup logger
import oemof
import pandas as pd
from glob import glob
from windnode_kwum.tools.logger import setup_logger
import pdb
from openpyxl import load_workbook
import openpyxl as xl
import os
from windnode_kwum.models.basic_model import create_model, simulate
from windnode_kwum.tools import config
#
import shutil
from win32com.client import DispatchEx


from windnode_kwum.tools.draw import draw_graph

# import oemof modules
from oemof import outputlib
from oemof.outputlib import processing, views
from oemof.graph import create_nx_graph

import matplotlib.pyplot as plt

def collect_specific_results(mainfolder, subfolders, kwum_path):

    sensi_name = subfolders
    folder = sensi_name

    results_path = mainfolder + sensi_name + '/'
    calculate_all_path = results_path + 'results/'
    results_filelist = [f for f in os.listdir(calculate_all_path) if  f.endswith("results.xlsx") if not f.startswith('~')]

    print(results_filelist)


    sheet_1 = 'Stromerzeugung'
    sheet_2 = 'Wärmenetze'
    sheet_3 = 'Anlagen'

    #writer_1 = pd.ExcelWriter(calculate_all_path + '!_' + sheet_1 + '_' + sensi_name + '.xlsx')
    #writer_2 = pd.ExcelWriter(calculate_all_path + '!_' + sheet_2 + '_' + sensi_name + '.xlsx')
    #writer_3 = pd.ExcelWriter(calculate_all_path + '!_' + sheet_3 + '_' + sensi_name + '.xlsx')

    ###########

    #master_sheet_1 = r'C:/Users/master/Google Drive/Masterarbeit/Auswertung/MASTER_' + sheet_1 + '.xlsx'
    #collection_sheet_1 = calculate_all_path + '!_' + sheet_1 + '_' + sensi_name + '.xlsx'
    #shutil.copy(master_sheet_1, collection_sheet_1)

    master_sheet_2 = kwum_path + 'Auswertung/MASTER_' + sheet_2 + '.xlsx'
    collection_sheet_2 = calculate_all_path + '!_' + sheet_2 + '_' + sensi_name + '.xlsx'
    shutil.copy(master_sheet_2, collection_sheet_2)

    #master_sheet_3 = r'C:/Users/master/Google Drive/Masterarbeit/Auswertung/MASTER_' + sheet_3 + '_2.xlsx'
    #collection_sheet_3 = calculate_all_path + '!_' + sheet_3 + '_' + sensi_name + '.xlsx'
    #shutil.copy(master_sheet_3, collection_sheet_3)

    ###########


    for f in results_filelist:
        f = f[:-5]
        print(f)
        s = f[:-8]
        results_file_path = calculate_all_path + f + '.xlsx'
    #    df_1 = pd.read_excel(results_file_path, sheet_1)
        df_2 = pd.read_excel(results_file_path, sheet_2)
    #    df_3 = pd.read_excel(results_file_path, sheet_3)

    #    book_1 = load_workbook(collection_sheet_1)
    #    writer_1 = pd.ExcelWriter(collection_sheet_1, engine='openpyxl')
    #    writer_1.book = book_1
    #    df_1.to_excel(writer_1, sheet_name=s)

        book_2 = load_workbook(collection_sheet_2)
        writer_2 = pd.ExcelWriter(collection_sheet_2, engine='openpyxl')
        writer_2.book = book_2
        df_2.to_excel(writer_2, sheet_name=s)

    #    book_3 = load_workbook(collection_sheet_3)
    #    writer_3 = pd.ExcelWriter(collection_sheet_3, engine='openpyxl')
    #    writer_3.book = book_3
    #    df_3.to_excel(writer_3, sheet_name=s)

    #    writer_1.save()
        writer_2.save()
    #    writer_3.save()

    #    writer_1.close()
        writer_2.close()
    #    writer_3.close()

    sensi_params_path = results_path + 'sensi_param_' + folder + '.xlsx'
    sensi_params_df = pd.read_excel(sensi_params_path)
    with pd.ExcelWriter(collection_sheet_2, engine='openpyxl') as writer:
        writer.book = load_workbook(collection_sheet_2)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        sensi_params_df.to_excel(writer, "sensi_params")

    excel = DispatchEx('Excel.Application')
    collection_sheet_refresh = excel.Workbooks.Open(collection_sheet_2)
    collection_sheet_refresh.RefreshAll()
    collection_sheet_refresh.Save()
    collection_sheet_refresh.Close()

    #with pd.ExcelWriter(collection_sheet_3, engine='openpyxl') as writer:
    #    writer.book = load_workbook(collection_sheet_3)
    #    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    #    sensi_params_df.to_excel(writer, "sensi_params")

    print("COLLECT RESULTS: DONE!")


def create_sensi_scenario_files(mainfolder, subfolder):
    calculate_all_path = mainfolder + subfolder + '/'

    # read sensi params in the folders
    df_sensi_param_all = pd.read_excel(calculate_all_path + 'sensi_param_' + subfolder + '.xlsx')
    print(df_sensi_param_all)
    df_sensi_param = df_sensi_param_all[df_sensi_param_all['filter'] == 1]
    df_sensi_param.index = df_sensi_param.index + 1
    print(df_sensi_param)
    # number of sensi scenario calculations
    ldf_num_series = df_sensi_param.lfd_Nr.astype(str)
    sensi_series_num = df_sensi_param.Nr.astype(str)
    # number each sceanrio with a number of 2 characters
    lfd_num = ldf_num_series.str.zfill(2)
    print("lfd_num")
    print(lfd_num)
    sensi_num = sensi_series_num.str.zfill(2)
    print("sensi_num")
    print(sensi_num)

    # copy the number of different scenario files for each sensi variation
    #sc_file_path = calculate_all_path + 'Master_2016_6c_sensi.xlsx'
    sc_master_file = [f for f in os.listdir(calculate_all_path) if f.startswith('MASTER') and f.endswith("scenario.xlsx")]
    sc_file_path = calculate_all_path + sc_master_file[0]

    for i, n in zip(lfd_num, sensi_num):
        copy_sc_file_path = calculate_all_path + subfolder + '_' + n + '.xlsx'
        shutil.copy(sc_file_path, copy_sc_file_path)

    return sensi_num


def run_scenario(cfg, esys):
    results = simulate(esys=esys,
                       solver=cfg['solver'])
    esys.results = results

    if cfg['dump']:
        path = os.path.join(config.get_data_root_dir(),
                            config.get('user_dirs',
                                       'results_dir')
                            )
        file = os.path.splitext(os.path.basename(__file__))[0] + '.oemof'

        esys.dump(dpath=path, filename=file)
        logger.info('The energy system was dumped to {}.'
                    .format(path + file))
    return esys, results

def export_results(esys, calculate_all_path, f):
    f = f
    results = esys.results
    busList = [item for item in esys.nodes if isinstance(item, oemof.solph.network.Bus)]
    busColorObject = {}
    results_flow_overview = pd.DataFrame()

    results_flows = []
    for bus in busList:
        # get bus from results
        bus_results = views.node(results, bus.label)
        bus_results_flows = bus_results['sequences']
        results_flows.append(bus_results_flows)

    results_flows = pd.concat(results_flows, axis=1)
    results_flows.to_excel(calculate_all_path + 'results/' + f + '_timeseries.xlsx')