import pandas as pd
from openpyxl import load_workbook
import openpyxl as xl
import shutil
from win32com.client import DispatchEx
#import win32com.client as win32
from windnode_kwum.scenarios.functions import collect_specific_results

import os

print("READ RESULTS SENSI")
# define main folder for calculation of several sensitivity analyses

kwum_path = os.getcwd().replace('WindNODE_KWUM\windnode_kwum\scenarios','')
main_folder_path = kwum_path + 'Scenarios\Scenario_folders/'
subfolders = os.listdir(main_folder_path)

# calculate every sensi analysis for each folder
for folder in subfolders:

    calculate_all_path = main_folder_path + folder + '/'
    szenario_workbook_path = kwum_path + 'Scenarios\scenario_data.xlsx'
    mainfolder = main_folder_path
    subfolders = folder

    # auswahlfilter aus sensi param datei
    '''
    # read sensi params in the folders
    df_sensi_param_all = pd.read_excel(calculate_all_path + 'sensi_param_' + subfolder + '.xlsx')
    print(df_sensi_param_all)
    df_sensi_param = df_sensi_param_all[df_sensi_param_all['filter'] == 1]
    df_sensi_param.index = df_sensi_param.index + 1
    print(df_sensi_param)
    # number of sensi scenario calculations
    ldf_num_series = df_sensi_param.lfd_Nr.astype(str)
    sensi_series_num = df_sensi_param.Nr.astype(str)
    # number each sceanrio with a number of 2 characters
    lfd_num = ldf_num_series.str.zfill(2)
    print("lfd_num")
    print(lfd_num)
    sensi_num = sensi_series_num.str.zfill(2)
    print("sensi_num")
    print(sensi_num)
    filelist = [subfolder + '_' + n + '.xlsx' for n in sensi_num]
    '''

    # ohne filter der sensi datei
    # define List of all scenario files for sensi analysis
    dirpath = os.getcwd()
    filelist = [f for f in os.listdir(calculate_all_path) if f.endswith('.xlsx') if not f.startswith('MASTER') if not f.startswith('sensi_param') if not f.startswith('~')]

    print(filelist)


    excel = DispatchEx('Excel.Application')
    scenario_workbook_file = excel.Workbooks.Open(szenario_workbook_path)
    sensi_param_file = excel.Workbooks.Open(calculate_all_path + 'sensi_param_' + folder + '.xlsx')
    sensi_param_file.RefreshAll()

    ts_df = []
    cost_flow_df = []
    for f in filelist:
        f = f[:-5]
        print(f)

        scenario_file = excel.Workbooks.Open(calculate_all_path + f + '.xlsx')
        scenario_file.RefreshAll()

        ts_path = calculate_all_path + 'results/' + f + '_timeseries.xlsx'
        print(ts_path)
        cost_flow_path = calculate_all_path + 'results/' + f + '_cost_flow.xlsx'

        ts_df = pd.read_excel(ts_path).round(5)
        ts_df_bool = ts_df.copy()
        ts_df_bool[ts_df_bool > 0] = 1
    #     ts_df.index = pd.to_datetime(ts_df[0])
    #     cost_flow_df = pd.read_excel(cost_flow_path)
    #     cost_flow_df.index = pd.to_datetime(cost_flow_df[0])

        flow_sum = ts_df.sum()
        flow_max = ts_df.max()
        flow_min = ts_df.min()
        flow_mean = ts_df.mean()

        flow_overview = pd.concat(
            [flow_sum.rename('sum'), flow_max.rename('max'), flow_min.rename('min'), flow_mean.rename('mean')], axis=1)

        # Version mit csv Datei
        #costs_revenues = pd.read_csv('C:/Users/master/Google Drive/Masterarbeit/KWUM_sensi_24_10/spot_prices.csv', sep=';')
        #costs_revenues.index = pd.to_datetime(costs_revenues.timestamp)

        # Version mit Sheet in Szenario Datei
        costs_revenues = pd.read_excel(calculate_all_path + f + '.xlsx', sheetname='costs_rev_ts')
        costs_revenues.index = pd.to_datetime(costs_revenues.timestamp)

        cost_flows = pd.DataFrame()
        cost_flows_rel = pd.DataFrame()

        component = 'chp_pr'
        flow = ((component, 'bus_chp_el'), 'flow')
        print(flow)
        if str(flow) in str(ts_df.columns.values):
            cost_flows[flow] = ts_df[str(flow)] * costs_revenues['base_revenues']
            cost_flows_rel[flow] = ts_df_bool[str(flow)] * costs_revenues['base_revenues']
            cost_flows_rel[flow] = cost_flows_rel[flow][cost_flows_rel[flow]!=0]

        component = 'chp_sch'
        flow = ((component, 'bus_chp_el'), 'flow')
        if str(flow) in str(ts_df.columns.values):
            cost_flows[flow] = ts_df[str(flow)] * costs_revenues['base_revenues']
            cost_flows_rel[flow] = ts_df_bool[str(flow)] * costs_revenues['base_revenues']
            cost_flows_rel[flow][cost_flows_rel[flow] != 0]
            cost_flows_rel[flow] = cost_flows_rel[flow][cost_flows_rel[flow] != 0]

        component = 'pth_pr'
        cs_flow = (('cs_to_' + component, 'bus_' + component), 'flow')
        flex_flow = (('flex_to_' + component, 'bus_' + component), 'flow')
        if str(cs_flow) in str(ts_df.columns.values):
            cost_flows[cs_flow] = ts_df[str(cs_flow)] * costs_revenues['spot_prices']
            cost_flows_rel[cs_flow] = ts_df_bool[str(cs_flow)] * costs_revenues['spot_prices']
            cost_flows_rel[cs_flow][cost_flows_rel[cs_flow] != 0]
            cost_flows_rel[cs_flow] = cost_flows_rel[cs_flow][cost_flows_rel[cs_flow] != 0]
        if str(flex_flow) in str(ts_df.columns.values):
            cost_flows[flex_flow] = ts_df[str(flex_flow)] * costs_revenues['flex_prices']
            cost_flows_rel[flex_flow] = ts_df_bool[str(flex_flow)] * costs_revenues['flex_prices']
            cost_flows_rel[flex_flow][cost_flows_rel[flex_flow] != 0]
            cost_flows_rel[flex_flow] = cost_flows_rel[flex_flow][cost_flows_rel[flex_flow] != 0]



        component = 'pth_sch'
        cs_flow = (('cs_to_' + component, 'bus_' + component), 'flow')
        flex_flow = (('flex_to_' + component, 'bus_' + component), 'flow')
        if str(cs_flow) in str(ts_df.columns.values):
            cost_flows[cs_flow] = ts_df[str(cs_flow)] * costs_revenues['spot_prices']
            cost_flows_rel[cs_flow] = ts_df_bool[str(cs_flow)] * costs_revenues['spot_prices']
            cost_flows_rel[cs_flow][cost_flows_rel[cs_flow] != 0]
            cost_flows_rel[cs_flow] = cost_flows_rel[cs_flow][cost_flows_rel[cs_flow] != 0]
        if str(flex_flow) in str(ts_df.columns.values):
            cost_flows[flex_flow] = ts_df[str(flex_flow)] * costs_revenues['flex_prices']
            cost_flows_rel[flex_flow] = ts_df_bool[str(flex_flow)] * costs_revenues['flex_prices']
            cost_flows_rel[flex_flow][cost_flows_rel[flex_flow] != 0]
            cost_flows_rel[flex_flow] = cost_flows_rel[flex_flow][cost_flows_rel[flex_flow] != 0]

        component = 'ptg_h2'
        cs_flow = (('cs_to_' + component, 'bus_' + component), 'flow')
        flex_flow = (('flex_to_' + component, 'bus_' + component), 'flow')
        if str(cs_flow) in str(ts_df.columns.values):
            cost_flows[cs_flow] = ts_df[str(cs_flow)] * costs_revenues['spot_prices']
            cost_flows_rel[cs_flow] = ts_df_bool[str(cs_flow)] * costs_revenues['spot_prices']
            cost_flows_rel[cs_flow][cost_flows_rel[cs_flow] != 0]
            cost_flows_rel[cs_flow] = cost_flows_rel[cs_flow][cost_flows_rel[cs_flow] != 0]
        if str(flex_flow) in str(ts_df.columns.values):
            cost_flows[flex_flow] = ts_df[str(flex_flow)] * costs_revenues['flex_prices']
            cost_flows_rel[flex_flow] = ts_df_bool[str(flex_flow)] * costs_revenues['flex_prices']
            cost_flows_rel[flex_flow][cost_flows_rel[flex_flow] != 0]
            cost_flows_rel[flex_flow] = cost_flows_rel[flex_flow][cost_flows_rel[flex_flow] != 0]

        component = 'ptg_pr'
        cs_flow = (('cs_to_' + component, 'bus_' + component), 'flow')
        flex_flow = (('flex_to_' + component, 'bus_' + component), 'flow')
        if str(cs_flow) in str(ts_df.columns.values):
            cost_flows[cs_flow] = ts_df[str(cs_flow)] * costs_revenues['spot_prices']
            cost_flows_rel[cs_flow] = ts_df_bool[str(cs_flow)] * costs_revenues['spot_prices']
            cost_flows_rel[cs_flow][cost_flows_rel[cs_flow] != 0]
            cost_flows_rel[cs_flow] = cost_flows_rel[cs_flow][cost_flows_rel[cs_flow] != 0]
        if str(flex_flow) in str(ts_df.columns.values):
            cost_flows[flex_flow] = ts_df[str(flex_flow)] * costs_revenues['flex_prices']
            cost_flows_rel[flex_flow] = ts_df_bool[str(flex_flow)] * costs_revenues['flex_prices']
            cost_flows_rel[flex_flow][cost_flows_rel[flex_flow] != 0]
            cost_flows_rel[flex_flow] = cost_flows_rel[flex_flow][cost_flows_rel[flex_flow] != 0]

        component = 'ptg_sch'
        cs_flow = (('cs_to_' + component, 'bus_' + component), 'flow')
        flex_flow = (('flex_to_' + component, 'bus_' + component), 'flow')
        if str(cs_flow) in str(ts_df.columns.values):
            cost_flows[cs_flow] = ts_df[str(cs_flow)] * costs_revenues['spot_prices']
            cost_flows_rel[cs_flow] = ts_df_bool[str(cs_flow)] * costs_revenues['spot_prices']
            cost_flows_rel[cs_flow][cost_flows_rel[cs_flow] != 0]
            cost_flows_rel[cs_flow] = cost_flows_rel[cs_flow][cost_flows_rel[cs_flow] != 0]
        if str(flex_flow) in str(ts_df.columns.values):
            cost_flows[flex_flow] = ts_df[str(flex_flow)] * costs_revenues['flex_prices']
            cost_flows_rel[flex_flow] = ts_df_bool[str(flex_flow)] * costs_revenues['flex_prices']
            cost_flows_rel[flex_flow][cost_flows_rel[flex_flow] != 0]
            cost_flows_rel[flex_flow] = cost_flows_rel[flex_flow][cost_flows_rel[flex_flow] != 0]



        component = 'batt_pr'
        cs_flow = (('cs_to_' + component, 'bus_' + component + '_in'), 'flow')
        cs_rev_flow = (('bus_' + component + '_out', component + '_to_cs'), 'flow')
        flex_flow = (('flex_to_' + component, 'bus_' + component + '_in'), 'flow')
        if str(cs_flow) in str(ts_df.columns.values):
            cost_flows[cs_flow] = ts_df[str(cs_flow)] * costs_revenues['spot_prices']
            cost_flows[cs_rev_flow] = ts_df[str(cs_rev_flow)] * costs_revenues['spot_revenues']
            cost_flows_rel[cs_flow] = ts_df_bool[str(cs_flow)] * costs_revenues['spot_prices']
            cost_flows_rel[cs_flow][cost_flows_rel[cs_flow] != 0]
            cost_flows_rel[cs_flow] = cost_flows_rel[cs_flow][cost_flows_rel[cs_flow] != 0]
            cost_flows_rel[cs_rev_flow] = ts_df_bool[str(cs_rev_flow)] * costs_revenues['spot_prices']
            cost_flows_rel[cs_rev_flow][cost_flows_rel[cs_rev_flow] != 0]
            cost_flows_rel[cs_rev_flow] = cost_flows_rel[cs_rev_flow][cost_flows_rel[cs_rev_flow] != 0]
        if str(flex_flow) in str(ts_df.columns.values):
            cost_flows[flex_flow] = ts_df[str(flex_flow)] * costs_revenues['flex_prices']
            cost_flows_rel[flex_flow] = ts_df_bool[str(flex_flow)] * costs_revenues['flex_prices']
            cost_flows_rel[flex_flow][cost_flows_rel[flex_flow] != 0]
            cost_flows_rel[flex_flow] = cost_flows_rel[flex_flow][cost_flows_rel[flex_flow] != 0]

        component = 'batt_sch'
        cs_flow = (('cs_to_' + component, 'bus_' + component + '_in'), 'flow')
        cs_rev_flow = (('bus_' + component + '_out', component + '_to_cs'), 'flow')
        flex_flow = (('flex_to_' + component, 'bus_' + component + '_in'), 'flow')
        if str(cs_flow) in str(ts_df.columns.values):
            cost_flows[cs_flow] = ts_df[str(cs_flow)] * costs_revenues['spot_prices']
            cost_flows[cs_rev_flow] = ts_df[str(cs_rev_flow)] * costs_revenues['spot_revenues']
            cost_flows_rel[cs_flow] = ts_df_bool[str(cs_flow)] * costs_revenues['spot_prices']
            cost_flows_rel[cs_flow][cost_flows_rel[cs_flow] != 0]
            cost_flows_rel[cs_flow] = cost_flows_rel[cs_flow][cost_flows_rel[cs_flow] != 0]
            cost_flows_rel[cs_rev_flow] = ts_df_bool[str(cs_rev_flow)] * costs_revenues['spot_prices']
            cost_flows_rel[cs_rev_flow][cost_flows_rel[cs_rev_flow] != 0]
            cost_flows_rel[cs_rev_flow] = cost_flows_rel[cs_rev_flow][cost_flows_rel[cs_rev_flow] != 0]
        if str(flex_flow) in str(ts_df.columns.values):
            cost_flows[flex_flow] = ts_df[str(flex_flow)] * costs_revenues['flex_prices']
            cost_flows_rel[flex_flow] = ts_df_bool[str(flex_flow)] * costs_revenues['flex_prices']
            cost_flows_rel[flex_flow][cost_flows_rel[flex_flow] != 0]
            cost_flows_rel[flex_flow] = cost_flows_rel[flex_flow][cost_flows_rel[flex_flow] != 0]

        cost_sum = cost_flows.sum()
        cost_max = cost_flows.max()
        cost_min = cost_flows.min()
        cost_mean = cost_flows.mean()
        cost_max_rel = cost_flows_rel.max()
        cost_min_rel = cost_flows_rel.min()
        cost_mean_rel = cost_flows_rel.mean()

        cost_overview = pd.concat(
            [cost_sum.rename('sum'), cost_max.rename('max'), cost_min.rename('min'), cost_mean.rename('mean'),
             cost_max_rel.rename('cost_rel_max'), cost_min_rel.rename('cost_rel_min'),
             cost_mean_rel.rename('cost_rel_mean')], axis=1)
        #

        # calculation_master_file = [f for f in os.listdir(calculate_all_path) if
        #                           f.startswith('MASTER') and f.endswith("results.xlsx")]
        # calculation_master_path = calculate_all_path + calculation_master_file[0]
        calculation_master_path = kwum_path + 'Scenarios/MASTER_cost_calculation.xlsx'

        diagram_master_path = kwum_path + 'Scenarios/MASTER_diagrams.xlsx'
        diagram_file_path = calculate_all_path + 'results/' + f + '_diagrams.xlsx'
        calculation_file_path = calculate_all_path + 'results/' + f + '_results.xlsx'
        print(calculation_master_path)
        print(calculation_file_path)
        shutil.copy(calculation_master_path, calculation_file_path)
        #    shutil.copy(diagram_master_path, diagram_file_path)

        with pd.ExcelWriter(calculation_file_path, engine='openpyxl') as writer:
            writer.book = xl.load_workbook(calculation_file_path)
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
            flow_overview.to_excel(writer, "PY_FLOW_RESULTS")
            cost_overview.to_excel(writer, "PY_COST_RESULTS")
            writer.save()
            writer.close()



        scenario_results_file = excel.Workbooks.Open(calculation_file_path)
        scenario_results_file.RefreshAll()
        scenario_results_file.Save()
        scenario_results_file.Close()

        #shutil.copy(calculation_file_path, calculate_all_path + 'results/' + f + '_results_var.xlsx')

        wb = xl.load_workbook(calculation_file_path, data_only=True)
        wb.save(calculation_file_path)


        scenario_file.Save()
        scenario_file.Close()


    sensi_param_file.Save()
    sensi_param_file.Close()
    scenario_workbook_file.RefreshAll()
    scenario_workbook_file.Save()
    scenario_workbook_file.Close()

    collect_specific_results(subfolders=folder, mainfolder=main_folder_path, kwum_path=kwum_path)



print("READ RESULTS SENSI: DONE!")


