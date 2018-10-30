from SALib.sample import saltelli
from SALib.analyze import sobol
import numpy as np
import os
from openpyxl import load_workbook
from windnode_kwum.scenarios import reference_scenario_curtailment
import csv


## THIS SENSITIVITY ANALYSIS IS BASED ON THE SOBOL ANALYSIS:
# https://salib.readthedocs.io/en/latest/api/SALib.analyze.html#module-SALib.analyze.sobol

## 1) Define the parameters to be varied:
#    Value to extract is the variable in which the impact of the inputs are measured: the energy flow of interest
#    x_param_to_be_varied: input parameters
#    y_param_to_be_varied: input parameters

value_to_extract = "bus_flex.P2H_sch"
x_param_to_be_varied = "transformers.P2H_sch.capacity"
y_param_to_be_varied = "storages.storage_th_sch.nominal capacity"

## 2) Define the problem:
#    num_vars: number of variables
#    names: name of the variable parameters
#    bounds: boundaries of our parameters

problem = {
  'num_vars': 2,
  'names': [x_param_to_be_varied, y_param_to_be_varied],
  'bounds': [[0, 68], [0, 680]]
}

# Saltelli sampling method: Generates model inputs using Saltelli's extension of the Sobol sequence.:
# https://salib.readthedocs.io/en/latest/_modules/SALib/sample/saltelli.html
X = saltelli.sample(problem, 10)
print('This is the value of X')
print(X)

Y = []

# Find the current file path for sensitivity_analysis.py
dirpath = os.getcwd()

dest = dirpath+'/data/reference_scenario_curtailment.xlsx'

csvData = [[y_param_to_be_varied, x_param_to_be_varied, "result"]]

for row in X:
  parameter_set = []
  data_from_run = []
  parameter_set.append({'key': x_param_to_be_varied, 'value': str(row[0])})
  parameter_set.append({'key': y_param_to_be_varied, 'value': str(row[1])})

  data_from_run.append(str(row[0]))
  data_from_run.append(str(row[1]))

## 3) OPEN, MODIFY AND SAVE EXCEL FILE #####

  # Open the reference_scenario_curtailment.xlsx for reading
  wb = load_workbook(filename=dest)

  for param in parameter_set:

    file_sections = param.get('key').split('.')

    # Open the sheet specified on the parameter_to_be_varied -> (tab.row.column)
    ws = wb[file_sections[0]]  # wb.get_sheet_by_name(file_sections[0])

    # Edit the cell specified on the parameter_to_be_varied -> (tab.row.column)
    row_to_edit = ''
    column_to_edit = ''

    for cell in ws.get_cell_collection():
      if cell.value == file_sections[1]:
        row_to_edit = cell.row

      if cell.value == file_sections[2]:
        column_to_edit = cell.col_idx

    ws.cell(row=row_to_edit, column=column_to_edit).value = param.get('value')

  # Save excel file with the changes made
  wb.save(dest)

  ### 4) EXECUTE reference_scenario_curtailment AND EXTRACT THE VALUE NEEDED from the energy optimization results:bus_results #####

  # Run the reference_scenario_curtailment.py to execute the  optimization of the energy system
  SA_variables = {
    "is_active": True,
    "results": data_from_run,
    "value_to_extract": value_to_extract
  }

  reference_scenario_curtailment.executeMain(SA_variables=SA_variables)

  Y.append(data_from_run[2])

  csvData.append(data_from_run)

with open(dirpath+'/results/sobol_sensitivity_results.csv', 'w') as csvFile:
    writer = csv.writer(csvFile)
    writer.writerows(csvData)

csvFile.close()

Y_numpy = np.array(Y, dtype=np.float)
print('This is the value of Y numpy')
print(Y_numpy)

# 5) SOBOL ANALYSIS
#    Perform analysis
Si = sobol.analyze(problem, Y_numpy, print_to_console=True, calc_second_order=True)

# Print the first-order sensitivity indices
# print(Si['S1'])
print('')
print('==================================')
print('Total-order index: contribution to the output variance caused by the model inputs')
print(Si['ST'])
print('==================================')
print('')
print('First order index')
print(Si['S1'])
print('==================================')
print('')
print('Second order index')
print(Si['S2'])
print('==================================')